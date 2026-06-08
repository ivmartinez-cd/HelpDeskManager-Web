import pandas as pd
import numpy as np
from datetime import date, datetime, timedelta
from pathlib import Path
import warnings
import io

warnings.filterwarnings("ignore")

# Tipos que provocan reset de contador
TIPOS_RESET = {"Reiniciar Contador"}


def cargar_archivo(ruta_o_stream) -> pd.DataFrame:
    try:
        df = pd.read_excel(ruta_o_stream, header=1)
    except Exception as e:
        raise ValueError(f"No se pudo abrir el archivo Excel. Asegúrate de que no esté dañado ni protegido.\nDetalle: {e}") from e

    df.columns = df.columns.str.strip()

    columnas_req = {"Articulo", "Nro Serie", "Sector", "Fecha", "Tipo Contador", "Clase", "Contador"}
    columnas_presentes = set(df.columns)
    faltantes = columnas_req - columnas_presentes
    if faltantes:
        raise ValueError(
            f"El archivo Excel no tiene el formato requerido.\n"
            f"Faltan las siguientes columnas: {', '.join(sorted(faltantes))}.\n\n"
            f"Columnas encontradas: {', '.join(sorted(df.columns))}"
        )

    df["Fecha"]    = pd.to_datetime(df["Fecha"], errors="coerce", dayfirst=True).dt.date
    df["Contador"] = pd.to_numeric(df["Contador"], errors="coerce")
    return df


def limpiar(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    mask = df["Fecha"].isna() | df["Contador"].isna()
    desc = df[mask].copy()
    desc["_motivo"] = "Fecha o Contador nulo/invalido"
    ok   = df[~mask].copy()
    ok["Contador"] = ok["Contador"].astype(int)
    return ok.reset_index(drop=True), desc.reset_index(drop=True)


def calcular_consumo_diario(lecturas: pd.DataFrame, min_dias_intervalo: int, umbral_minimo_consumo: float) -> tuple[float, str, list[str]]:
    notas = []
    if len(lecturas) < 2:
        return 0.0, "Una sola lectura — sin tendencia calculable", notas

    deltas = []
    for i in range(1, len(lecturas)):
        prev = lecturas.iloc[i - 1]
        curr = lecturas.iloc[i]
        dias = (curr["Fecha"] - prev["Fecha"]).days
        diff = int(curr["Contador"]) - int(prev["Contador"])

        if dias < min_dias_intervalo:
            notas.append(
                f"Intervalo {prev['Fecha']}->{curr['Fecha']}: {dias}d — omitido (mismo dia)"
            )
            continue
        if diff < 0:
            notas.append(
                f"Intervalo {prev['Fecha']}->{curr['Fecha']}: diff={diff} — omitido (reset o crecimiento negativo)"
            )
            continue

        deltas.append(diff / dias)

    if not deltas:
        return 0.0, "Sin intervalos validos (todos negativos o mismo dia)", notas

    consumo = float(np.mean(deltas))
    if consumo <= umbral_minimo_consumo:
        notas.append(
            f"Consumo promedio calculado ({consumo:.4f}) por debajo o igual al umbral ({umbral_minimo_consumo}) -> forzado a 0.0"
        )
        consumo = 0.0
        desc = f"Consumo por debajo del umbral de {umbral_minimo_consumo}"
    else:
        desc = f"Promedio de {len(deltas)} intervalo(s) valido(s)"
    return consumo, desc, notas


def procesar_combinacion(
    serie: str,
    clase: str,
    articulo: str,
    sector: str,
    df_comb: pd.DataFrame,
    fecha_toma: date,
    tolerancia_dias: int,
    min_dias_intervalo: int,
    ventana_reciente_dias: int,
    umbral_minimo_consumo: float,
    max_antiguedad_lectura_dias: int,
) -> tuple[dict, list[dict]]:

    auditoria = []
    df_comb   = df_comb.sort_values("Fecha").reset_index(drop=True)

    df_real   = df_comb[df_comb["Fecha"] == fecha_toma]
    df_hist   = df_comb[df_comb["Fecha"]  < fecha_toma].reset_index(drop=True)
    df_futuro = df_comb[df_comb["Fecha"]  > fecha_toma]

    # Registrar lecturas futuras (solo validacion, no usadas en proyeccion)
    for _, row in df_futuro.iterrows():
        auditoria.append(_aud(serie, clase, row, usado=False,
            motivo=f"Lectura futura (>{fecha_toma}) — reservada para validacion, no usada en proyeccion"))

    # ── REAL: existe lectura exactamente en fecha_toma ──
    if not df_real.empty:
        ultima = df_real.sort_values("Fecha").iloc[-1]
        for _, row in df_real.iterrows():
            es = (row["Fecha"] == ultima["Fecha"] and row["Contador"] == ultima["Contador"])
            auditoria.append(_aud(serie, clase, row, usado=es,
                motivo="Lectura real en fecha de toma — utilizada" if es
                        else "Lectura real en fecha de toma — existe otra mas reciente"))
        for _, row in df_hist.iterrows():
            auditoria.append(_aud(serie, clase, row, usado=False,
                motivo="Lectura historica no requerida — existe lectura real en fecha de toma"))

        return _resultado(serie, clase, articulo, sector,
            fecha_lectura  = ultima["Fecha"],
            contador_base  = int(ultima["Contador"]),
            dias_proy      = 0,
            consumo_diario = 0.0,
            paginas_sumadas = 0,
            fecha_toma     = fecha_toma,
            contador_proy  = int(ultima["Contador"]),
            metodo         = "REAL",
            obs            = f"Lectura real del {ultima['Fecha']} (tipo: {ultima['Tipo Contador']})",
        ), auditoria

    # ── PROYECTADO: no hay lectura en fecha_toma ──
    if df_hist.empty:
        auditoria.append({
            "Nro Serie": serie, "Clase": clase,
            "Fecha": None, "Contador": None, "Tipo Contador": None,
            "Usado": False,
            "Motivo": "Sin lecturas historicas disponibles",
        })
        return _resultado(serie, clase, articulo, sector,
            fecha_lectura=None, contador_base=None, dias_proy=None,
            consumo_diario=None, paginas_sumadas=None, fecha_toma=fecha_toma, contador_proy=None,
            metodo="SIN DATOS",
            obs="No existen lecturas previas a la fecha de toma",
        ), auditoria

    ultima_hist = df_hist.iloc[-1]
    dias_proy   = (fecha_toma - ultima_hist["Fecha"]).days

    # ── Lectura suficientemente reciente: usar como REAL sin proyectar ──
    if dias_proy <= tolerancia_dias:
        for _, row in df_hist.iterrows():
            es_ultima = (row["Fecha"] == ultima_hist["Fecha"] and row["Contador"] == ultima_hist["Contador"])
            auditoria.append(_aud(serie, clase, row, usado=es_ultima,
                motivo=f"Lectura a {dias_proy} dia(s) de la fecha de toma — usada como REAL sin proyectar"
                       if es_ultima else "Lectura historica no requerida — existe lectura reciente suficiente"))
        return _resultado(serie, clase, articulo, sector,
            fecha_lectura  = ultima_hist["Fecha"],
            contador_base  = int(ultima_hist["Contador"]),
            dias_proy      = 0,
            consumo_diario = 0.0,
            paginas_sumadas = 0,
            fecha_toma     = fecha_toma,
            contador_proy  = int(ultima_hist["Contador"]),
            metodo         = "REAL",
            obs            = f"Lectura del {ultima_hist['Fecha']} a {dias_proy} dia(s) de la fecha de toma — sin proyeccion",
        ), auditoria

    # Encontrar el último reset en df_hist
    idx_ultimo_reset = -1
    for i in range(len(df_hist)):
        row = df_hist.iloc[i]
        cond1 = row["Tipo Contador"] in TIPOS_RESET
        cond2 = False
        if i > 0:
            prev_row = df_hist.iloc[i - 1]
            if row["Contador"] < prev_row["Contador"]:
                cond2 = True
        if cond1 or cond2:
            idx_ultimo_reset = i

    # Separar df_hist en descartados y válidos
    fecha_reset = None
    if idx_ultimo_reset != -1:
        fecha_reset = df_hist.iloc[idx_ultimo_reset]["Fecha"]
        df_valid = df_hist.iloc[idx_ultimo_reset:]
    else:
        df_valid = df_hist

    # Aplicar filtro de ventana reciente de N días respecto a la última lectura válida
    ultima_fecha = df_valid.iloc[-1]["Fecha"]
    fecha_limite_reciente = ultima_fecha - timedelta(days=ventana_reciente_dias)
    df_reciente = df_valid[df_valid["Fecha"] >= fecha_limite_reciente]

    usar_solo_reciente = False
    if len(df_reciente) >= 2:
        usar_solo_reciente = True
        df_final_calculo = df_reciente
    else:
        df_final_calculo = df_valid

    for idx_hist, row in df_hist.iterrows():
        # Caso 1: Anterior al último reset
        if idx_ultimo_reset != -1 and idx_hist < idx_ultimo_reset:
            auditoria.append(_aud(serie, clase, row, usado=False,
                motivo=f"Lectura historica descartada — anterior al ultimo reset del {fecha_reset}"))
        else:
            # Caso 2: Post-reset pero fuera de la ventana reciente (si se usa la ventana reciente)
            fuera_de_ventana = (usar_solo_reciente and row["Fecha"] < fecha_limite_reciente)
            if fuera_de_ventana:
                auditoria.append(_aud(serie, clase, row, usado=False,
                    motivo=f"Lectura historica descartada — fuera de la ventana reciente de {ventana_reciente_dias} dias (limite: {fecha_limite_reciente})"))
            else:
                # Caso 3: Usada en la tendencia
                is_reset_point = (idx_ultimo_reset != -1 and idx_hist == idx_ultimo_reset)
                if is_reset_point:
                    motivo_reset = ""
                    if row["Tipo Contador"] in TIPOS_RESET:
                        motivo_reset = f" [RESET: Tipo Contador '{row['Tipo Contador']}']"
                    else:
                        prev_val = int(df_hist.iloc[idx_ultimo_reset - 1]["Contador"])
                        curr_val = int(row["Contador"])
                        motivo_reset = f" [RESET: Contador disminuyo de {prev_val} a {curr_val}]"
                    auditoria.append(_aud(serie, clase, row, usado=True,
                        motivo=f"Lectura historica — inicio de tendencia tras reset{motivo_reset}"))
                else:
                    is_first_reciente = (usar_solo_reciente and idx_hist == df_reciente.index[0])
                    if is_first_reciente:
                        auditoria.append(_aud(serie, clase, row, usado=True,
                            motivo=f"Lectura historica — inicio de tendencia (ventana reciente de {ventana_reciente_dias} dias)"))
                    else:
                        auditoria.append(_aud(serie, clase, row, usado=True,
                            motivo="Lectura historica — usada para calcular tendencia"))

    consumo, desc_consumo, notas_calc = calcular_consumo_diario(df_final_calculo, min_dias_intervalo, umbral_minimo_consumo)

    for nota in notas_calc:
        auditoria.append({
            "Nro Serie": serie, "Clase": clase,
            "Fecha": None, "Contador": None, "Tipo Contador": None,
            "Usado": False,
            "Motivo": f"[Tendencia] {nota}",
        })

    if dias_proy > max_antiguedad_lectura_dias:
        auditoria.append({
            "Nro Serie": serie, "Clase": clase,
            "Fecha": None, "Contador": None, "Tipo Contador": None,
            "Usado": False,
            "Motivo": f"[Tendencia] Ultima lectura disponible ({ultima_hist['Fecha']}) supera la antiguedad maxima permitida ({max_antiguedad_lectura_dias} dias) -> tendencia forzada a 0.0",
        })
        consumo = 0.0
        desc_consumo = f"Lectura base muy antigua (> {max_antiguedad_lectura_dias} dias)"

    contador_proy = int(round(ultima_hist["Contador"] + consumo * dias_proy))

    obs_parts = [desc_consumo, f"{dias_proy} dia(s) proyectado(s) hasta {fecha_toma}"]
    if usar_solo_reciente:
        obs_parts.insert(0, f"Tendencia calculada sobre ventana reciente de {ventana_reciente_dias} dias")
    if consumo == 0.0:
        obs_parts.append("Consumo = 0 — se replica el ultimo contador conocido")

    paginas_sumadas = contador_proy - int(ultima_hist["Contador"])

    return _resultado(serie, clase, articulo, sector,
        fecha_lectura  = ultima_hist["Fecha"],
        contador_base  = int(ultima_hist["Contador"]),
        dias_proy      = dias_proy,
        consumo_diario = round(consumo, 4),
        paginas_sumadas = paginas_sumadas,
        fecha_toma     = fecha_toma,
        contador_proy  = contador_proy,
        metodo         = "PROYECTADO",
        obs            = " | ".join(obs_parts),
    ), auditoria


def _aud(serie, clase, row, usado, motivo) -> dict:
    return {
        "Nro Serie": serie, "Clase": clase,
        "Fecha": row["Fecha"], "Contador": row["Contador"],
        "Tipo Contador": row["Tipo Contador"],
        "Usado": usado, "Motivo": motivo,
    }


def _resultado(
    serie, clase, articulo, sector,
    fecha_lectura, contador_base, dias_proy,
    consumo_diario, paginas_sumadas, fecha_toma, contador_proy, metodo, obs,
) -> dict:
    return {
        "Nro Serie"               : serie,
        "Clase"                   : clase,
        "Articulo"                : articulo,
        "Sector"                  : sector,
        "Fecha Ultima Lectura"    : fecha_lectura,
        "Contador Base"           : contador_base,
        "Dias Proyectados"        : dias_proy,
        "Consumo Diario Promedio" : consumo_diario,
        "Paginas Sumadas"         : paginas_sumadas,
        "Fecha Toma"              : fecha_toma,
        "Contador Proyectado"     : contador_proy,
        "Metodo"                  : metodo,
        "Observaciones"           : obs,
    }


def construir_validacion(df_res: pd.DataFrame, df_aud: pd.DataFrame, fecha_toma: date) -> pd.DataFrame:
    if df_aud.empty:
        return pd.DataFrame(columns=[
            "Nro Serie", "Clase", "Fecha Toma", "Contador Proyectado",
            "Fecha Lectura Real", "Contador Real",
            "Diferencia", "Error %", "Nota",
        ])

    futuras = df_aud[
        df_aud["Motivo"].str.startswith("Lectura futura", na=False) &
        df_aud["Contador"].notna()
    ].copy()

    if futuras.empty:
        return pd.DataFrame(columns=[
            "Nro Serie", "Clase", "Fecha Toma", "Contador Proyectado",
            "Fecha Lectura Real", "Contador Real",
            "Diferencia", "Error %", "Nota",
        ])

    proy = df_res[df_res["Metodo"] == "PROYECTADO"][
        ["Nro Serie", "Clase", "Fecha Toma", "Contador Proyectado"]
    ]

    futuras = futuras.sort_values("Fecha")
    primera_futura = futuras.groupby(["Nro Serie", "Clase"]).first().reset_index()

    val = proy.merge(
        primera_futura[["Nro Serie", "Clase", "Fecha", "Contador"]].rename(
            columns={"Fecha": "Fecha Lectura Real", "Contador": "Contador Real"}
        ),
        on=["Nro Serie", "Clase"], how="left",
    )

    mask = val["Contador Real"].notna()
    if not val.empty and mask.any():
        val.loc[mask, "Diferencia"] = (
            val.loc[mask, "Contador Proyectado"] - val.loc[mask, "Contador Real"]
        ).astype(int)
        val.loc[mask, "Error %"] = (
            (val.loc[mask, "Diferencia"] / val.loc[mask, "Contador Real"]) * 100
        ).round(2)

    val["Nota"] = val["Contador Real"].apply(
        lambda x: "Lectura real posterior disponible" if pd.notna(x)
                  else "Sin lectura real posterior aun"
    )

    return val[[
        "Nro Serie", "Clase", "Fecha Toma", "Contador Proyectado",
        "Fecha Lectura Real", "Contador Real",
        "Diferencia", "Error %", "Nota",
    ]]


def construir_siges(df_res: pd.DataFrame, fecha_toma: date) -> pd.DataFrame:
    # Solo series que tienen al menos un contador PROYECTADO
    series_proy = df_res[df_res["Metodo"] == "PROYECTADO"]["Nro Serie"].unique()
    df_filtrado = df_res[df_res["Nro Serie"].isin(series_proy)]

    filas = []

    for serie, grupo in df_filtrado.groupby("Nro Serie", sort=False):
        mono  = grupo[grupo["Clase"] == "Mono" ].iloc[0] if (grupo["Clase"] == "Mono" ).any() else None
        color = grupo[grupo["Clase"] == "Color"].iloc[0] if (grupo["Clase"] == "Color").any() else None

        ref   = mono if mono is not None else color
        fecha = ref["Fecha Toma"] if ref is not None else fecha_toma

        obs_parts = []
        if mono is not None and pd.notna(mono["Observaciones"]):
            obs_parts.append(f"Mono: {mono['Observaciones']}")
        if color is not None and pd.notna(color["Observaciones"]):
            obs_parts.append(f"Color: {color['Observaciones']}")
        if mono is not None and color is not None and \
           mono["Observaciones"] == color["Observaciones"]:
            obs_parts = [str(mono["Observaciones"])]
        observaciones = " | ".join(obs_parts)

        metodos = grupo["Metodo"].unique().tolist()
        motivo  = metodos[0] if len(metodos) == 1 else " / ".join(metodos)

        filas.append({
            "SERIE"        : serie,
            "FECHA"        : fecha,
            "TIPO"         : 4,
            "CLASE_10"     : 10 if mono  is not None else "",
            "CONTADOR_10"  : int(mono ["Contador Proyectado"]) if mono  is not None and pd.notna(mono ["Contador Proyectado"]) else "",
            "CLASE_20"     : 20 if color is not None else "",
            "CONTADOR_20"  : int(color["Contador Proyectado"]) if color is not None and pd.notna(color["Contador Proyectado"]) else "",
            "MOTIVO"       : motivo,
            "OBSERVACIONES": observaciones,
        })

    if not filas:
        return pd.DataFrame(columns=[
            "SERIE", "FECHA", "TIPO", "CLASE_10", "CONTADOR_10", "CLASE_20", "CONTADOR_20", "MOTIVO", "OBSERVACIONES"
        ])

    df = pd.DataFrame(filas)
    for col in ("TIPO", "CLASE_10", "CONTADOR_10", "CLASE_20", "CONTADOR_20"):
        df[col] = pd.array(df[col].replace("", pd.NA), dtype="Int64")
    for col in ("MOTIVO", "OBSERVACIONES"):
        df[col] = df[col].fillna("")
    return df


def _escribir_kpis(writer, stats: dict):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = writer.book
    ws = wb.create_sheet("KPIs", 0)

    C_AZUL_TITULO    = "1F4E79"
    C_AZUL_MEDIO     = "2E75B6"
    C_VERDE          = "375623"
    C_VERDE_CLARO    = "E2EFDA"
    C_NARANJA        = "833C00"
    C_NARANJA_CLARO  = "FCE4D6"
    C_GRIS           = "595959"
    C_GRIS_CLARO     = "F2F2F2"
    C_ROJO           = "C00000"
    C_ROJO_CLARO     = "FFDFD7"
    C_AZUL_CLARO     = "DEEAF1"
    C_AMARILLO       = "7F6000"
    C_AMARILLO_CLARO = "FFF2CC"
    C_BLANCO         = "FFFFFF"

    borde = Border(
        left  =Side(style="thin", color="BFBFBF"),
        right =Side(style="thin", color="BFBFBF"),
        top   =Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )

    def fill(h): return PatternFill("solid", fgColor=h)
    def font(bold=False, size=10, color="000000"):
        return Font(name="Calibri", bold=bold, size=size, color=color)
    def aln(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    def merge(r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    def cell(r, c, val="", fnt=None, fll=None, aln_=None, brd=None):
        cl = ws.cell(r, c)
        cl.value = val
        if fnt:  cl.font      = fnt
        if fll:  cl.fill      = fll
        if aln_: cl.alignment = aln_
        if brd:  cl.border    = brd
        return cl

    SEP_COLS  = [1, 5, 9, 13, 17, 21]
    CARD_COLS = [2, 6, 10, 14, 18]
    CARD_W    = 4

    for c in range(1, 23):
        ws.column_dimensions[get_column_letter(c)].width = (1 if c in SEP_COLS else 8)

    def kpi_card(row, col_start, etiqueta, valor, sub, fondo, c_valor, c_label):
        c2 = col_start + CARD_W - 1
        for r in range(row, row + 4):
            for c in range(col_start, col_start + CARD_W):
                ws.cell(r, c).fill   = fill(fondo)
                ws.cell(r, c).border = borde
        merge(row,     col_start, row,     c2)
        merge(row + 1, col_start, row + 1, c2)
        cell(row + 1, col_start, etiqueta,
             fnt=font(bold=False, size=9, color=c_label),
             aln_=aln("center", "center"))
        merge(row + 2, col_start, row + 2, c2)
        cell(row + 2, col_start, valor,
             fnt=font(bold=True, size=22, color=c_valor),
             aln_=aln("center", "center"))
        merge(row + 3, col_start, row + 3, c2)
        cell(row + 3, col_start, sub,
             fnt=font(bold=False, size=9, color=c_label),
             aln_=aln("center", "center"))

    ws.row_dimensions[1].height  = 28
    ws.row_dimensions[2].height  = 6
    for r in [3]:     ws.row_dimensions[r].height = 8
    for r in [4]:     ws.row_dimensions[r].height = 14
    for r in [5]:     ws.row_dimensions[r].height = 36
    for r in [6]:     ws.row_dimensions[r].height = 14
    ws.row_dimensions[7].height  = 8
    for r in [8]:     ws.row_dimensions[r].height = 8
    for r in [9]:     ws.row_dimensions[r].height = 14
    for r in [10]:    ws.row_dimensions[r].height = 36
    for r in [11]:    ws.row_dimensions[r].height = 14
    ws.row_dimensions[12].height = 10

    # Título
    merge(1, 1, 1, 21)
    cell(1, 1,
         f"PROYECCION DE CONTADORES  |  Fecha de toma: {stats['fecha_toma']}  |  {stats['archivo']}",
         fnt=font(bold=True, size=13, color=C_BLANCO),
         fll=fill(C_AZUL_TITULO),
         aln_=aln("center", "center"))

    # Fila 1 de KPIs (filas 3-6)
    total  = stats["total"]
    reales = stats["reales"]
    proy   = stats["proyectados"]
    sin_d  = stats["sin_datos"]

    kpi_card(3, CARD_COLS[0], "TOTAL PROCESADOS",
             f"{total:,}", "Serie x Clase",
             C_AZUL_CLARO, C_AZUL_TITULO, C_AZUL_MEDIO)

    kpi_card(3, CARD_COLS[1], "REAL",
             f"{reales:,}", f"{reales/total*100:.1f}% del total" if total else "—",
             C_VERDE_CLARO, C_VERDE, C_VERDE)

    kpi_card(3, CARD_COLS[2], "PROYECTADO",
             f"{proy:,}", f"{proy/total*100:.1f}% del total" if total else "—",
             C_NARANJA_CLARO, C_NARANJA, C_NARANJA)

    kpi_card(3, CARD_COLS[3], "SIN DATOS",
             f"{sin_d:,}", f"{sin_d/total*100:.1f}% del total" if total else "—",
             C_ROJO_CLARO, C_ROJO, C_ROJO)

    kpi_card(3, CARD_COLS[4], "LECTURAS HISTORICAS",
             f"{stats['n_hist']:,}", "usadas para proyectar",
             C_GRIS_CLARO, C_GRIS, C_GRIS)

    # Fila 2 de KPIs (filas 8-11)
    kpi_card(8, CARD_COLS[0], "DIAS PROYECT. MEDIANA",
             f"{stats['dias_mediana']:.0f}", f"max: {stats['dias_max']} dias",
             C_AMARILLO_CLARO, C_AMARILLO, C_AMARILLO)

    kpi_card(8, CARD_COLS[1], "ALTA PRECISION  (1 dia)",
             f"{stats['n_1dia']:,}", f"{stats['pct_1dia']:.1f}% de proyectados",
             C_VERDE_CLARO, C_VERDE, C_VERDE)

    kpi_card(8, CARD_COLS[2], "CONSUMO MEDIANA",
             f"{stats['consumo_mediana']:.1f}", "pag/dia (proyectados)",
             C_AZUL_CLARO, C_AZUL_TITULO, C_AZUL_MEDIO)

    kpi_card(8, CARD_COLS[3], "LECTURAS FUTURAS",
             f"{stats['n_futuro']:,}", "para validacion posterior",
             C_GRIS_CLARO, C_GRIS, C_GRIS)

    kpi_card(8, CARD_COLS[4], "CONSUMO MAX",
             f"{stats['consumo_max']:.0f}", "pag/dia",
             C_NARANJA_CLARO, C_NARANJA, C_NARANJA)

    # Tabla distribución días proyectados
    ROW_TABLA = 13
    COLS_TABLA = [2, 6, 10, 14]
    MERGE_INTERP_HASTA = 20

    ws.row_dimensions[ROW_TABLA].height = 20
    merge(ROW_TABLA, 2, ROW_TABLA, 20)
    cell(ROW_TABLA, 2,
         "DISTRIBUCION POR DISTANCIA A FECHA DE TOMA  (series proyectadas)",
         fnt=font(bold=True, size=11, color=C_BLANCO),
         fll=fill(C_AZUL_MEDIO),
         aln_=aln("center", "center"))

    f = ROW_TABLA + 1
    ws.row_dimensions[f].height = 16
    for col, enc in zip(COLS_TABLA, ["Rango", "Cantidad", "% proyectados", "Interpretacion"]):
        es_interp = enc == "Interpretacion"
        c2 = MERGE_INTERP_HASTA if es_interp else col + CARD_W - 2
        merge(f, col, f, c2)
        cell(f, col, enc,
             fnt=font(bold=True, size=10, color=C_BLANCO),
             fll=fill(C_AZUL_TITULO),
             aln_=aln("center", "center"),
             brd=borde)
    f += 1

    dist_rows = [
        ("1 dia",         stats["dist"]["1 dia"],       "Maxima precision — base solida a 1 dia de la fecha de toma"),
        ("2 - 7 dias",    stats["dist"]["2-7 dias"],    "Alta precision — tendencia reciente muy representativa"),
        ("8 - 30 dias",   stats["dist"]["8-30 dias"],   "Buena precision — variaciones menores esperables"),
        ("31 - 90 dias",  stats["dist"]["31-90 dias"],  "Precision media — el consumo puede haber variado"),
        ("91 - 365 dias", stats["dist"]["91-365 dias"], "Estimacion aproximada — mayor margen de error"),
        ("+365 dias",     stats["dist"]["+365 dias"],   "Referencia historica — proyeccion orientativa"),
    ]
    for i, (rango, cant, interp) in enumerate(dist_rows):
        fondo = C_BLANCO if i % 2 == 0 else C_GRIS_CLARO
        pct   = f"{cant/proy*100:.1f}%" if proy else "—"
        ws.row_dimensions[f].height = 15
        for col, val, es_interp in [
            (COLS_TABLA[0], rango,  False),
            (COLS_TABLA[1], cant,   False),
            (COLS_TABLA[2], pct,    False),
            (COLS_TABLA[3], interp, True),
        ]:
            c2 = MERGE_INTERP_HASTA if es_interp else col + CARD_W - 2
            merge(f, col, f, c2)
            cell(f, col, val,
                 fnt=font(bold=(not es_interp and val == rango), size=10),
                 fll=fill(fondo),
                 aln_=aln("center" if not es_interp else "left", "center"),
                 brd=borde)
        f += 1


def _escribir_leyenda(writer):
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = writer.book
    ws = wb.create_sheet("Leyenda")

    azul_oscuro  = PatternFill("solid", fgColor="1F4E79")
    azul_medio   = PatternFill("solid", fgColor="2E75B6")
    gris_claro   = PatternFill("solid", fgColor="F2F2F2")
    blanco       = PatternFill("solid", fgColor="FFFFFF")
    fuente_titulo = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    fuente_seccion= Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    fuente_campo  = Font(name="Calibri", bold=True, size=10)
    fuente_normal = Font(name="Calibri", size=10)
    wrap          = Alignment(wrap_text=True, vertical="top")

    def titulo(fila, texto):
        ws.merge_cells(f"A{fila}:D{fila}")
        c = ws.cell(fila, 1, texto)
        c.font = fuente_titulo
        c.fill = azul_oscuro
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[fila].height = 20

    def seccion(fila, texto):
        ws.merge_cells(f"A{fila}:D{fila}")
        c = ws.cell(fila, 1, texto)
        c.font = fuente_seccion
        c.fill = azul_medio
        c.alignment = Alignment(vertical="center")
        ws.row_dimensions[fila].height = 18

    def fila_dato(fila, campo, descripcion, es_par=False):
        relleno = gris_claro if es_par else blanco
        c1 = ws.cell(fila, 1, campo)
        c1.font = fuente_campo
        c1.fill = relleno
        c1.alignment = wrap
        ws.merge_cells(f"B{fila}:D{fila}")
        c2 = ws.cell(fila, 2, descripcion)
        c2.font = fuente_normal
        c2.fill = relleno
        c2.alignment = wrap
        ws.row_dimensions[fila].height = max(15, min(len(descripcion) // 6, 60))

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 28

    f = 1
    titulo(f, "LEYENDA — Proyeccion de Contadores"); f += 1

    seccion(f, "COLUMNAS — Hoja 'Proyeccion'"); f += 1
    campos_proy = [
        ("Nro Serie", "Numero de serie del equipo (identificador unico de hardware)."),
        ("Clase", "Tipo de impresion: MONO o COLOR. Cada equipo puede tener un contador Mono y uno Color independiente."),
        ("Articulo", "Modelo o descripcion comercial del equipo."),
        ("Sector", "Ubicacion fisica del equipo dentro del cliente."),
        ("Fecha Ultima Lectura", "Fecha de la lectura mas reciente utilizada como base."),
        ("Contador Base", "Valor del contador en la Fecha Ultima Lectura."),
        ("Dias Proyectados", "Cantidad de dias entre la Fecha Ultima Lectura y la Fecha de Toma."),
        ("Consumo Diario Promedio", "Paginas impresas por dia en promedio."),
        ("Paginas Sumadas", "Paginas estimadas sumadas (Consumo Diario x Dias Proyectados)."),
        ("Fecha Toma", "Fecha objetivo de la proyeccion."),
        ("Contador Proyectado", "Valor estimado del contador en la Fecha de Toma."),
        ("Metodo", "REAL: lectura exacta. PROYECTADO: estimacion. SIN DATOS: sin lecturas previas."),
        ("Observaciones", "Detalle tecnico del calculo."),
    ]
    for i, (campo, desc) in enumerate(campos_proy):
        fila_dato(f, campo, desc, i % 2 == 0); f += 1

    f += 1

    seccion(f, "COMO SE CALCULA EL CONSUMO DIARIO PROMEDIO"); f += 1
    pasos = [
        ("Paso 1 — Historial", "Se toman TODAS las lecturas del equipo anteriores a la Fecha de Toma."),
        ("Paso 2 — Intervalos", "Se forman pares de lecturas consecutivas."),
        ("Paso 3 — Consumo por intervalo", "(Contador final - Contador inicial) / Cantidad de dias."),
        ("Paso 4 — Descartes", "Se descartan intervalos si el contador disminuyo o si la diferencia de dias es 0."),
        ("Paso 5 — Promedio final", "Se promedian los consumos diarios de todos los intervalos validos."),
    ]
    for i, (campo, desc) in enumerate(pasos):
        fila_dato(f, campo, desc, i % 2 == 0); f += 1


def ejecutar_proyeccion(
    file_input,
    fecha_toma: date,
    folder_salida: str,
    tolerancia_dias: int = 2,
    min_dias_intervalo: int = 1,
    ventana_reciente_dias: int = 365,
    umbral_minimo_consumo: float = 0.2,
    max_antiguedad_lectura_dias: int = 365,
) -> tuple[str, str, list[str], dict, list[dict], list[dict]]:
    """
    Ejecuta el algoritmo de proyecciones a partir de un archivo Excel en memoria o disco.
    Retorna (ruta_excel, ruta_csv, logs, kpi_stats, registros, validacion).
    """
    logs = []
    
    # 1. Cargar y limpiar
    df_raw = cargar_archivo(file_input)
    df, desc_limpieza = limpiar(df_raw)

    if df.empty:
        raise ValueError("El archivo Excel no contiene registros válidos para procesar (todas las filas tienen Fecha o Contador nulos).")

    n_real   = (df["Fecha"] == fecha_toma).sum()
    n_hist   = (df["Fecha"]  < fecha_toma).sum()
    n_futuro = (df["Fecha"]  > fecha_toma).sum()

    if n_hist == 0:
        logs.append("No existen lecturas históricas anteriores a la fecha de toma. Todas las combinaciones serán SIN DATOS.")

    combos = df.groupby(["Nro Serie", "Clase"])
    info = (
        df.sort_values("Fecha")
          .groupby(["Nro Serie", "Clase"])[["Articulo", "Sector"]]
          .last().reset_index()
    )

    resultados      = []
    auditoria_total = []

    for (serie, clase), df_comb in combos:
        row_info = info[(info["Nro Serie"] == serie) & (info["Clase"] == clase)]
        articulo = row_info["Articulo"].values[0] if not row_info.empty else ""
        sector   = row_info["Sector"].values[0]   if not row_info.empty else ""
        res, aud = procesar_combinacion(
            str(serie), str(clase), articulo, sector, df_comb,
            fecha_toma=fecha_toma,
            tolerancia_dias=tolerancia_dias,
            min_dias_intervalo=min_dias_intervalo,
            ventana_reciente_dias=ventana_reciente_dias,
            umbral_minimo_consumo=umbral_minimo_consumo,
            max_antiguedad_lectura_dias=max_antiguedad_lectura_dias
        )
        resultados.append(res)
        auditoria_total.extend(aud)

    df_res = pd.DataFrame(resultados)
    df_aud = pd.DataFrame(auditoria_total)
    df_val = construir_validacion(df_res, df_aud, fecha_toma)
    df_siges = construir_siges(df_res, fecha_toma)

    total      = len(df_res)
    reales     = (df_res["Metodo"] == "REAL").sum()
    proyectados= (df_res["Metodo"] == "PROYECTADO").sum()
    sin_datos  = (df_res["Metodo"] == "SIN DATOS").sum()

    df_proy = df_res[df_res["Metodo"] == "PROYECTADO"]["Dias Proyectados"].dropna()
    cons_proy = df_res[df_res["Metodo"] == "PROYECTADO"]["Consumo Diario Promedio"].dropna()
    cons_proy_pos = cons_proy[cons_proy > 0]

    # Distribución de días
    bins   = [0, 1, 7, 30, 90, 365, 9999]
    labels = ["1 dia", "2-7 dias", "8-30 dias", "31-90 dias", "91-365 dias", "+365 dias"]
    cats   = pd.cut(df_proy, bins=bins, labels=labels, right=True)
    dist   = {lbl: int(cats.value_counts().get(lbl, 0)) for lbl in labels}

    kpi_stats = {
        "fecha_toma"     : str(fecha_toma),
        "archivo"        : "Excel Subido",
        "total"          : total,
        "reales"         : int(reales),
        "proyectados"    : int(proyectados),
        "sin_datos"      : int(sin_datos),
        "n_hist"         : int(n_hist),
        "n_futuro"       : int(n_futuro),
        "n_1dia"         : int(dist["1 dia"]),
        "pct_1dia"       : dist["1 dia"] / proyectados * 100 if proyectados else 0,
        "dias_mediana"   : float(df_proy.median()) if not df_proy.empty else 0,
        "dias_max"       : int(df_proy.max())      if not df_proy.empty else 0,
        "consumo_mediana": float(cons_proy_pos.median()) if not cons_proy_pos.empty else 0,
        "consumo_max"    : float(cons_proy_pos.max())    if not cons_proy_pos.empty else 0,
        "dist"           : dist,
    }

    # Definir rutas de salida
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_excel = f"Reporte_Proyeccion_{stamp}.xlsx"
    nombre_csv = f"import_siges_proyeccion_{stamp}.csv"

    ruta_excel = Path(folder_salida) / nombre_excel
    ruta_csv = Path(folder_salida) / nombre_csv

    # Escribir Excel
    with pd.ExcelWriter(ruta_excel, engine="openpyxl") as writer:
        df_res.to_excel(writer, sheet_name="Proyeccion",     index=False)
        df_siges.to_excel(writer, sheet_name="CSV para SiGes", index=False)
        df_val.to_excel(writer, sheet_name="Validacion",    index=False)
        df_aud.to_excel(writer, sheet_name="Auditoria",     index=False)

        resumen_rows = [
            ("Fecha de toma",             str(fecha_toma)),
            ("Total registros",           len(df_raw)),
            ("Lecturas en fecha de toma", int(n_real)),
            ("Lecturas historicas",       int(n_hist)),
            ("Lecturas futuras",          int(n_futuro)),
            ("Combinaciones procesadas",  int(total)),
            ("REAL",                      int(reales)),
            ("PROYECTADO",                int(proyectados)),
            ("SIN DATOS",                 int(sin_datos)),
        ]
        if not df_proy.empty:
            resumen_rows += [
                ("Dias proyectados — min",  int(df_proy.min())),
                ("Dias proyectados — media",round(float(df_proy.mean()),1)),
                ("Dias proyectados — max",  int(df_proy.max())),
            ]
        pd.DataFrame(resumen_rows, columns=["Indicador","Valor"]).to_excel(
            writer, sheet_name="Resumen", index=False
        )

        if not desc_limpieza.empty:
            desc_limpieza.to_excel(writer, sheet_name="Descartados_Limpieza", index=False)

        _escribir_kpis(writer, kpi_stats)
        _escribir_leyenda(writer)

        for sheet_name in writer.sheets:
            if sheet_name == "Leyenda":
                continue
            ws = writer.sheets[sheet_name]
            for col in ws.columns:
                try:
                    max_len = max((
                        9 if isinstance(c.value, bool) and c.value else (
                            5 if isinstance(c.value, bool) and not c.value else len(str(c.value))
                        )
                        for c in col if c.value is not None
                    ), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
                except AttributeError:
                    pass

    # Escribir CSV para SiGes
    df_siges.to_csv(
        ruta_csv, sep=";", index=False, encoding="utf-8", lineterminator="\r\n"
    )

    def _serialize(df: pd.DataFrame) -> list[dict]:
        out = []
        for row in df.to_dict("records"):
            clean = {}
            for k, v in row.items():
                if hasattr(v, "isoformat"):
                    clean[k] = v.isoformat()
                elif isinstance(v, float) and v != v:  # NaN
                    clean[k] = None
                else:
                    clean[k] = v
            out.append(clean)
        return out

    return str(ruta_excel), str(ruta_csv), logs, kpi_stats, _serialize(df_res), _serialize(df_val)
